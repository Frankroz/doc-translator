import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.cell.cell import MergedCell
from deep_translator import GoogleTranslator
import copy
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import os
import platform

try:
    import win32com.client as win32
except ImportError:
    win32 = None

class BatchTranslatorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Traductor Pro Excel XLS/XLSX - V8")
        self.root.geometry("750x600")
        self.root.configure(bg="#f4f4f4")

        self.is_cancelled = False
        self.translator = GoogleTranslator(source='auto', target='es')
        self.file_vars = []

        # --- UI SETUP ---
        self.header = tk.Label(root, text="Traductor Multiformato de Excel", 
                              font=("Segoe UI", 16, "bold"), bg="#f4f4f4")
        self.header.pack(pady=15)

        frame_top = tk.Frame(root, bg="#f4f4f4")
        frame_top.pack(pady=5)

        self.btn_folder = tk.Button(frame_top, text="📂 Seleccionar Carpeta", command=self.select_folder, 
                                   width=22, bg="#4CAF50", fg="white", font=("Segoe UI", 10, "bold"))
        self.btn_folder.pack(side=tk.LEFT, padx=10)

        self.btn_file = tk.Button(frame_top, text="📄 Seleccionar Archivo", command=self.select_single_file, 
                                 width=22, bg="#2196F3", fg="white", font=("Segoe UI", 10, "bold"))
        self.btn_file.pack(side=tk.LEFT, padx=10)

        self.container = tk.Frame(root, bd=1, relief="sunken")
        self.container.pack(fill="both", expand=True, padx=30, pady=10)

        self.canvas = tk.Canvas(self.container, bg="white", highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.container, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg="white")

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.bottom_panel = tk.Frame(root, bg="#e0e0e0")
        self.bottom_panel.pack(fill="x", side="bottom")

        self.status_label = tk.Label(self.bottom_panel, text="Esperando archivos...", bg="#e0e0e0", font=("Segoe UI", 10, "bold"))
        self.status_label.pack(pady=5)

        btn_frame = tk.Frame(self.bottom_panel, bg="#e0e0e0")
        btn_frame.pack(pady=5)

        self.btn_start = tk.Button(btn_frame, text="🚀 INICIAR", command=self.start_batch_thread, 
                                  state="disabled", width=15, bg="#FF9800", fg="white", font=("Segoe UI", 10, "bold"))
        self.btn_start.pack(side=tk.LEFT, padx=5)

        self.btn_cancel = tk.Button(btn_frame, text="🛑 CANCELAR", command=self.cancel_process, 
                                   state="disabled", width=15, bg="#f44336", fg="white", font=("Segoe UI", 10, "bold"))
        self.btn_cancel.pack(side=tk.LEFT, padx=5)

        self.progress = ttk.Progressbar(self.bottom_panel, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", pady=(10, 0))

    def _on_mousewheel(self, event):
        scale = -1 if platform.system() == 'Windows' else 1
        self.canvas.yview_scroll(int(scale*(event.delta/120)), "units")

    def cancel_process(self):
        if messagebox.askyesno("Cancelar", "¿Deseas detener la traducción actual?"):
            self.is_cancelled = True

    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.clear_list()
            files = [f for f in os.listdir(folder) if (f.lower().endswith('.xlsx') or f.lower().endswith('.xls')) and not f.startswith('~$')]
            for f in files: self.add_item(os.path.join(folder, f))
            if self.file_vars: self.btn_start.config(state="normal")

    def select_single_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.clear_list()
            self.add_item(path)
            self.btn_start.config(state="normal")

    def add_item(self, path):
        var = tk.BooleanVar(value=True)
        chk = tk.Checkbutton(self.scrollable_frame, text=os.path.basename(path), variable=var, bg="white", anchor="w")
        chk.pack(fill="x", padx=10, pady=2)
        self.file_vars.append((path, var))

    def clear_list(self):
        for widget in self.scrollable_frame.winfo_children(): widget.destroy()
        self.file_vars = []
        self.btn_start.config(state="disabled")

    def toggle_ui(self, enable):
        state = "normal" if enable else "disabled"
        self.btn_start.config(state=state)
        self.btn_folder.config(state=state)
        self.btn_file.config(state=state)
        self.btn_cancel.config(state="disabled" if enable else "normal")

    def start_batch_thread(self):
        selected = [p for p, v in self.file_vars if v.get()]
        if not selected: return
        self.is_cancelled = False
        self.toggle_ui(False)
        threading.Thread(target=self.run_batch, args=(selected,), daemon=True).start()

    def convert_xls_to_xlsx_win32(self, xls_path):
        if win32 is None:
            raise Exception("Librería 'pywin32' no instalada.")
            
        abs_path = os.path.abspath(xls_path)
        temp_xlsx = abs_path + "_temp.xlsx"
        
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        
        try:
            wb = excel.Workbooks.Open(abs_path)
            wb.SaveAs(temp_xlsx, FileFormat=51)
            wb.Close()
            return temp_xlsx
        finally:
            excel.Quit()

    def run_batch(self, files):
        total = len(files)
        for i, path in enumerate(files):
            if self.is_cancelled: break
            fname = os.path.basename(path)
            self.status_label.config(text=f"Traduciendo ({i+1}/{total}): {fname}", fg="#d32f2f")
            
            try:
                self.process_excel(path)
            except Exception as e:
                messagebox.showerror("Error", f"Error en {fname}:\n{str(e)}")
            
            self.progress['value'] = ((i + 1) / total) * 100
            self.root.update_idletasks()

        self.toggle_ui(True)
        self.progress['value'] = 0
        self.status_label.config(text="Proceso finalizado", fg="green")
        messagebox.showinfo("Fin", "La tarea ha concluido.")

    def process_excel(self, input_file):
        is_old_format = input_file.lower().endswith('.xls')
        path_to_load = input_file

        if is_old_format:
            path_to_load = self.convert_xls_to_xlsx_win32(input_file)

        try:
            # data_only=False para mantener fórmulas si las hay
            wb = openpyxl.load_workbook(path_to_load)
            new_wb = openpyxl.Workbook()

            for idx, sname in enumerate(wb.sheetnames):
                if self.is_cancelled: break
                ws = wb[sname]
                new_ws = new_wb.active if idx == 0 else new_wb.create_sheet(sname)
                new_ws.title = sname

                # COPIA DE DIMENSIONES MEJORADA
                for c, d in ws.column_dimensions.items(): new_ws.column_dimensions[c].width = d.width
                for r, d in ws.row_dimensions.items(): new_ws.row_dimensions[r].height = d.height

                # DETECCIÓN DE RANGO REAL (Evita cortes a la mitad)
                max_row = ws.max_row
                max_col = ws.max_column

                for r in range(1, max_row + 1):
                    if self.is_cancelled: break
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        new_cell = new_ws.cell(row=r, column=c)

                        if not isinstance(cell, MergedCell):
                            if cell.value and isinstance(cell.value, str):
                                try:
                                    t = self.translator.translate(cell.value)
                                    new_cell.value = f"{cell.value}\n({t})"
                                except:
                                    new_cell.value = cell.value
                            else:
                                new_cell.value = cell.value

                        # COPIA DE ESTILOS EXPLÍCITA
                        if cell.has_style:
                            new_cell.font = copy.copy(cell.font)
                            new_cell.border = copy.copy(cell.border)
                            new_cell.fill = copy.copy(cell.fill)
                            new_cell.alignment = Alignment(
                                wrap_text=True, 
                                horizontal=cell.alignment.horizontal or 'center', 
                                vertical=cell.alignment.vertical or 'center'
                            )

                # COPIA DE MERGES
                for mr in ws.merged_cells.ranges:
                    new_ws.merge_cells(str(mr))

                # COPIA DE IMÁGENES
                if hasattr(ws, '_images'):
                    for img in ws._images:
                        new_ws.add_image(copy.copy(img))

            output_name = f"BILINGUAL_{os.path.splitext(os.path.basename(input_file))[0]}.xlsx"
            output_file = os.path.join(os.path.dirname(input_file), output_name)
            new_wb.save(output_file)
        
        finally:
            if is_old_format and os.path.exists(path_to_load):
                wb.close()
                os.remove(path_to_load)

if __name__ == "__main__":
    root = tk.Tk()
    app = BatchTranslatorGUI(root)
    root.mainloop()